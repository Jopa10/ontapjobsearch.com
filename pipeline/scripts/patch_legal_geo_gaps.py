from __future__ import annotations

from pathlib import Path

from openpyxl import load_workbook

PATH = Path("pipeline/geo/geo_lookup.xlsx")

# One-shot canonical corrections found while validating the Legal family across all 78 markets.
AREA_FIXES = {
    "diss": ("Diss", "Norfolk"),
    "studley": ("Studley", "West Midlands - Coventry & Warwickshire"),
}

FALLBACK_FIXES = {
    "sheffield": ("Sheffield", "Yorkshire - South"),
    "leeds": ("Leeds", "Yorkshire - West"),
    "belfast": ("Belfast", "Northern Ireland - East"),
}


def norm(value: object) -> str:
    return str(value or "").strip().casefold()


def ensure_area_sheet(ws) -> None:
    headers = {str(cell.value or "").strip(): idx for idx, cell in enumerate(ws[1], 1)}
    area_col = headers["Area"]
    cluster_col = headers["Cluster"]
    seen: set[str] = set()
    for row in range(2, ws.max_row + 1):
        key = norm(ws.cell(row, area_col).value)
        if key in AREA_FIXES:
            area, cluster = AREA_FIXES[key]
            ws.cell(row, area_col).value = area
            ws.cell(row, cluster_col).value = cluster
            seen.add(key)
    for key, (area, cluster) in AREA_FIXES.items():
        if key not in seen:
            ws.append([area if idx == area_col else cluster if idx == cluster_col else "" for idx in range(1, ws.max_column + 1)])


def ensure_fallback_sheet(ws) -> None:
    headers = {str(cell.value or "").strip(): idx for idx, cell in enumerate(ws[1], 1)}
    status_col = headers["Status"]
    location_col = headers["Location"]
    cluster_col = headers["Cluster"]
    seen: set[str] = set()
    for row in range(2, ws.max_row + 1):
        key = norm(ws.cell(row, location_col).value)
        if key in FALLBACK_FIXES:
            location, cluster = FALLBACK_FIXES[key]
            ws.cell(row, status_col).value = "auto"
            ws.cell(row, location_col).value = location
            ws.cell(row, cluster_col).value = cluster
            seen.add(key)
    for key, (location, cluster) in FALLBACK_FIXES.items():
        if key not in seen:
            values = [""] * ws.max_column
            values[status_col - 1] = "auto"
            values[location_col - 1] = location
            values[cluster_col - 1] = cluster
            ws.append(values)


def main() -> None:
    wb = load_workbook(PATH)
    ensure_area_sheet(wb[wb.sheetnames[0]])
    ensure_fallback_sheet(wb["LocationFallback"])
    wb.save(PATH)

    check = load_workbook(PATH, data_only=True)
    main_ws = check[check.sheetnames[0]]
    fallback_ws = check["LocationFallback"]
    area_rows = {norm(r[0]): norm(r[1]) for r in main_ws.iter_rows(min_row=2, values_only=True) if len(r) >= 2}
    fh = {str(c.value or "").strip(): i for i, c in enumerate(fallback_ws[1])}
    fallback_rows = {
        norm(r[fh["Location"]]): norm(r[fh["Cluster"]])
        for r in fallback_ws.iter_rows(min_row=2, values_only=True)
        if norm(r[fh["Status"]]) == "auto"
    }
    expected_area = {k: norm(v[1]) for k, v in AREA_FIXES.items()}
    expected_fallback = {k: norm(v[1]) for k, v in FALLBACK_FIXES.items()}
    assert all(area_rows.get(k) == v for k, v in expected_area.items()), (area_rows, expected_area)
    assert all(fallback_rows.get(k) == v for k, v in expected_fallback.items()), (fallback_rows, expected_fallback)
    print("Canonical Legal geo gaps patched and verified")


if __name__ == "__main__":
    main()

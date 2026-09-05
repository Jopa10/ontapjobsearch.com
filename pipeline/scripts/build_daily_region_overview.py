from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass
from datetime import datetime
import csv
import io
import json
from pathlib import Path
import re
import subprocess
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

try:
    from .live_job_source_counter import LiveInventory, collect_live_inventory
except ImportError:
    from live_job_source_counter import LiveInventory, collect_live_inventory


PIPELINE_ROOT = Path(__file__).resolve().parents[1]
REPO_ROOT = PIPELINE_ROOT.parent
CATALOG = PIPELINE_ROOT / "config" / "uk_assessable_regions.json"
REGISTER = PIPELINE_ROOT / "registers" / "region_category_slice_register.csv"
OUTPUT = PIPELINE_ROOT / "reports-daily" / "daily-region-overview.md"
JOBG8_CATEGORY_PROFILE = PIPELINE_ROOT / "reports-daily" / "jobg8-feed-category-profile.csv"
GEO_LOOKUP = PIPELINE_ROOT / "geo" / "geo_lookup.xlsx"
CITY_PAGE_REGISTER = PIPELINE_ROOT / "city_pages" / "city-page-register.json"
BROAD_CITY_PAGE_REGISTER = PIPELINE_ROOT / "city_pages" / "broad-city-page-register.json"
CITY_PAGE_THRESHOLD = 4
NON_LOCALITY_LABELS = {
    "city",
    "not specified",
    "remote",
    "various",
    "multiple locations",
    "nationwide",
    "uk",
    "united kingdom",
}
LOCALITY_ALIASES = {
    "newcastle upon tyne": "newcastle",
    "newcastle-upon-tyne": "newcastle",
    "brighton": "brighton & hove",
}
EXPECTED_REGION_COUNT = 78
CORE_ROUTES = {
    "/",
    "/browse-jobs",
    "/about",
    "/ai-tips",
    "/contact",
    "/privacy-policy",
    "/terms-of-service",
    "/sector-switching",
}

FAMILIES = (
    {
        "key": "service_admin",
        "label": "Service admin",
        "register_category": "admin_service",
        "source_category": "Admin/Service – Office Support",
        "decision_report": "pipeline/reports-daily/decision-report-admin-service.csv",
        "profile_category": "admin_service",
        "candidate_dir": "output-admin-service",
        "candidate_pattern": "{slug}-admin-service.json",
        "published_filenames": ("service-administrator-jobs.json",),
    },
    {
        "key": "support_worker",
        "label": "Support worker",
        "register_category": "support_worker",
        "source_category": "Support Worker – Wide",
        "decision_report": "pipeline/reports-daily/decision-report-support-worker.csv",
        "profile_category": "support_worker",
        "candidate_dir": "output-support-worker",
        "candidate_pattern": "{slug}-support-worker.json",
        "published_filenames": ("support-worker.json", "support-worker-jobs.json"),
    },
    {
        "key": "sales_advisor",
        "label": "Sales advisor",
        "register_category": "customer_sales",
        "source_category": "Customer Sales / Sales Advisor",
        "decision_report": "",
        "profile_category": "",
        "candidate_dir": "",
        "candidate_pattern": "",
        "published_slug": "customer-sales-jobs",
        "published_filenames": ("customer-sales-jobs.json",),
    },
    {
        "key": "legal_assistant_paralegal",
        "label": "Paralegal",
        "register_category": "legal_assistant_paralegal",
        "source_category": "",
        "decision_report": "",
        "profile_category": "",
        "candidate_dir": "",
        "candidate_pattern": "",
        "published_slug": "paralegal-jobs",
        "published_filenames": ("paralegal-jobs.json",),
    },
    {
        "key": "marketing",
        "label": "Marketing",
        "register_category": "marketing",
        "source_category": "",
        "decision_report": "",
        "profile_category": "",
        "candidate_dir": "",
        "candidate_pattern": "",
        "published_slug": "marketing-jobs",
        "published_filenames": ("marketing-jobs.json",),
    },
    {
        "key": "finance_accounts",
        "label": "Finance / Accounts",
        "register_category": "finance_accounts",
        "source_category": "",
        "decision_report": "",
        "profile_category": "",
        "candidate_dir": "",
        "candidate_pattern": "",
        "published_slug": "finance-accounts-jobs",
        "published_filenames": ("finance-accounts-jobs.json",),
    },
    {
        "key": "hr_recruitment",
        "label": "HR / Recruitment",
        "register_category": "hr_recruitment",
        "source_category": "",
        "decision_report": "",
        "profile_category": "",
        "candidate_dir": "",
        "candidate_pattern": "",
        "published_slug": "hr-recruitment-jobs",
        "published_filenames": ("hr-recruitment-jobs.json",),
    },
    {
        "key": "customer_service_contact_centre",
        "label": "CS / Contact centre",
        "register_category": "customer_service_contact_centre",
        "source_category": "Customer Service / Contact Centre",
        "decision_report": "",
        "profile_category": "customer_service_contact_centre",
        "candidate_dir": "",
        "candidate_pattern": "",
        "published_slug": "customer-service-jobs",
        "published_filenames": ("customer-service-jobs.json",),
    },
)

FAMILY_BY_FILENAME = {
    filename: family["key"]
    for family in FAMILIES
    for filename in family["published_filenames"]
}


@dataclass(frozen=True)
class SourceReportSummary:
    path: str
    report_date: str
    total_live_jobs: int


@dataclass(frozen=True)
class SiteInventorySummary:
    report_date: str
    unique_live_jobs: int
    unique_jobg8_jobs: int
    unique_external_jobs: int
    provider_counts: dict[str, int]
    provider_duplicate_jobs: dict[str, int]
    provider_extra_placements: dict[str, int]
    slice_counts: dict[tuple[str, str], int]
    slice_placements: int
    jobs_on_multiple_slices: int
    extra_slice_placements: int
    jobs_outside_governed_slices: int
    non_live_slice_jobs: int


@dataclass(frozen=True)
class JobG8CategoryProfile:
    feed_date: str
    total_jobs: int
    published_jobs: int | None
    counts: tuple[tuple[str, int, int | None], ...]


@dataclass(frozen=True)
class CityOpportunityRow:
    status: str
    locality: str
    region: str
    live_jobs: int
    existing_pages: int
    routes: str


def _normalise_place(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "")).strip().casefold()


def _location_candidates(value: str) -> tuple[str, ...]:
    """Return conservative exact-place candidates from a published location."""
    clean = re.sub(
        r"\s*\((?:[^)]*\b(?:hybrid|remote|home[- ]based|working)\b[^)]*)\)\s*$",
        "",
        value,
        flags=re.IGNORECASE,
    ).strip()
    candidates = [clean]
    if "," in clean:
        candidates.append(clean.split(",", 1)[0].strip())
    return tuple(dict.fromkeys(item for item in candidates if item))


def _load_mapped_localities(path: Path = GEO_LOOKUP) -> dict[str, tuple[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook["Sheet1"]
        rows = sheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header != ("Area", "Cluster"):
            raise RuntimeError(f"Unexpected geo lookup header in {path}: {header}")
        result: dict[str, tuple[str, str]] = {}
        for area, cluster in rows:
            area_text = str(area or "").strip()
            cluster_text = str(cluster or "").strip()
            key = _normalise_place(area_text)
            if key and cluster_text:
                result[key] = (area_text, cluster_text)
        return result
    finally:
        workbook.close()


def _active_city_pages(
    path: Path = CITY_PAGE_REGISTER,
    broad_path: Path | None = None,
) -> dict[str, list[str]]:
    raw = _load_json(path)
    broad_raw = _load_json(BROAD_CITY_PAGE_REGISTER) if broad_path is None and path == CITY_PAGE_REGISTER else []
    if broad_path is not None:
        broad_raw = _load_json(broad_path)
    if not isinstance(raw, list) or not isinstance(broad_raw, list):
        raise RuntimeError("City-page registers must be arrays")
    result: dict[str, list[str]] = defaultdict(list)
    for item in [*raw, *broad_raw]:
        if not isinstance(item, dict) or _normalise_place(item.get("lifecycle_state")) != "active":
            continue
        locality = _normalise_place(item.get("display_name"))
        route = str(item.get("route") or "").strip()
        if locality and route:
            result[locality].append(route)
    return result


def _city_opportunity_rows(
    inventory: LiveInventory,
    *,
    geo_lookup_path: Path = GEO_LOOKUP,
    city_register_path: Path = CITY_PAGE_REGISTER,
) -> tuple[list[CityOpportunityRow], int, int]:
    """Aggregate every unique live job by its recognised stated town/locality."""
    mapped = _load_mapped_localities(geo_lookup_path)
    active_pages = _active_city_pages(city_register_path)
    counts: Counter[str] = Counter()
    labels: dict[str, tuple[str, str]] = {}
    mapped_jobs = 0

    for job in inventory.jobs:
        raw_location = _normalise_place(job.location)
        if raw_location in NON_LOCALITY_LABELS:
            continue
        match: tuple[str, str] | None = None
        for candidate in _location_candidates(job.location):
            candidate_key = _normalise_place(candidate)
            candidate_key = LOCALITY_ALIASES.get(candidate_key, candidate_key)
            if candidate_key in active_pages:
                mapped_match = mapped.get(candidate_key)
                match = (
                    mapped_match[0] if mapped_match else candidate_key.title(),
                    mapped_match[1] if mapped_match else job.region,
                )
            else:
                match = mapped.get(candidate_key)
            if match:
                break
        if not match:
            continue
        locality, region = match
        if _normalise_place(locality) == _normalise_place(region):
            continue
        key = _normalise_place(locality)
        counts[key] += 1
        labels[key] = (locality, region)
        mapped_jobs += 1

    keys = set(counts) | set(active_pages)
    rows: list[CityOpportunityRow] = []
    for key in keys:
        locality, region = labels.get(key, (key.title(), ""))
        routes = sorted(set(active_pages.get(key, [])))
        count = counts.get(key, 0)
        if routes:
            status = "LIVE PAGE"
        elif _normalise_place(region) == "london":
            status = "HOLD – LONDON"
        elif count >= CITY_PAGE_THRESHOLD:
            status = "CREATE"
        else:
            status = "MONITOR"
        rows.append(
            CityOpportunityRow(
                status=status,
                locality=locality,
                region=region,
                live_jobs=count,
                existing_pages=len(routes),
                routes=", ".join(routes),
            )
        )

    order = {"CREATE": 0, "LIVE PAGE": 1, "HOLD – LONDON": 2, "MONITOR": 3}
    rows.sort(key=lambda row: (order[row.status], -row.live_jobs, row.locality.casefold()))
    return rows, mapped_jobs, len(inventory.jobs) - mapped_jobs


def _load_jobg8_category_profile() -> JobG8CategoryProfile | None:
    if not JOBG8_CATEGORY_PROFILE.is_file():
        return None
    with JOBG8_CATEGORY_PROFILE.open("r", encoding="utf-8-sig", newline="") as handle:
        rows = list(csv.DictReader(handle))
    required = {"feed_date", "total_jobs", "jobg8_category", "count"}
    if not rows or not required.issubset(rows[0]):
        raise RuntimeError(f"Unexpected JobG8 category profile: {JOBG8_CATEGORY_PROFILE}")
    feed_dates = {(row.get("feed_date") or "").strip() for row in rows}
    totals = {int((row.get("total_jobs") or "0").strip()) for row in rows}
    if len(feed_dates) != 1 or len(totals) != 1:
        raise RuntimeError("JobG8 category profile contains inconsistent feed dates or totals")
    has_published_counts = {"published_jobg8_jobs", "published_count"}.issubset(rows[0])
    published_totals = {
        int((row.get("published_jobg8_jobs") or "0").strip()) for row in rows
    } if has_published_counts else set()
    if has_published_counts and len(published_totals) != 1:
        raise RuntimeError("JobG8 category profile contains inconsistent published totals")
    counts = tuple(
        (
            (row.get("jobg8_category") or "(blank)").strip(),
            int((row.get("count") or "0").strip()),
            int((row.get("published_count") or "0").strip()) if has_published_counts else None,
        )
        for row in rows
    )
    total = totals.pop()
    if sum(count for _category, count, _published in counts) != total:
        raise RuntimeError("JobG8 category profile does not reconcile to its stated total")
    published_total = published_totals.pop() if has_published_counts else None
    if published_total is not None and sum(count or 0 for _category, _received, count in counts) != published_total:
        raise RuntimeError("JobG8 category profile does not reconcile to its published total")
    return JobG8CategoryProfile(feed_dates.pop(), total, published_total, counts)


def _load_family_coverage_date() -> str:
    path = PIPELINE_ROOT / "reports-daily" / "daily-family-coverage.csv"
    with path.open("r", encoding="utf-8-sig", newline="") as handle:
        dates = {(row.get("feed_date") or "").strip() for row in csv.DictReader(handle)}
    if len(dates) != 1 or not next(iter(dates)):
        raise RuntimeError(f"Expected one current family-coverage feed date, found {dates}")
    return dates.pop()


def _load_json(path: Path):
    if not path.is_file():
        return None
    try:
        return json.loads(path.read_text(encoding="utf-8-sig"))
    except (json.JSONDecodeError, OSError):
        return None


def _count_json_data(data) -> int:
    if isinstance(data, list):
        return len(data)
    if isinstance(data, dict):
        for key in ("jobs", "items", "results"):
            value = data.get(key)
            if isinstance(value, list):
                return len(value)
    return 0


def _job_count(path: Path) -> int:
    return _count_json_data(_load_json(path))


def _published_configured_count(region_slug: str, category_slug: str) -> int:
    path = REPO_ROOT / "app" / "_city-pages" / "configured-slices" / region_slug / f"{category_slug}.json"
    return _job_count(path) if path.is_file() else 0


def _load_statuses() -> dict[tuple[str, str], str]:
    statuses: dict[tuple[str, str], str] = {}
    with REGISTER.open("r", encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames != ["region", "category", "status"]:
            raise RuntimeError(f"Unexpected slice register header: {reader.fieldnames}")
        for row in reader:
            region = (row.get("region") or "").strip()
            category = (row.get("category") or "").strip()
            status = (row.get("status") or "").strip().upper()
            if region and category:
                statuses[(region, category)] = status
    return statuses


def _published_page_inventory() -> tuple[list[list[object]], dict[str, int]]:
    """Mirror the public sitemap's non-job routes and published job-page gate."""
    sitemap_source = (REPO_ROOT / "app" / "sitemap.ts").read_text(encoding="utf-8")
    base_match = re.search(r"const baseRoutes = \[(.*?)\]", sitemap_source, re.S)
    if not base_match:
        raise RuntimeError("Could not read baseRoutes from app/sitemap.ts")
    base_routes = re.findall(r"['\"](/[^'\"]*)['\"]", base_match.group(1))

    route_catalog = _load_json(PIPELINE_ROOT / "config" / "job_slice_catalog.json") or {}
    categories = route_catalog.get("categories", {})
    regions = route_catalog.get("regions", {})
    statuses = _load_statuses()

    details: list[list[object]] = []
    seen_routes: set[str] = set()

    def add(page_type: str, area: str, family: str, route: str, jobs: int | str = "") -> None:
        if route in seen_routes:
            return
        seen_routes.add(route)
        details.append(["Detail", page_type, area, family, route, 1, jobs, "Yes"])

    for route in base_routes:
        if route in CORE_ROUTES:
            add("Core", "Sitewide", "", route)
            continue
        parts = route.strip("/").split("/")
        area = parts[0].replace("-", " ").title() if parts else ""
        family = parts[-1].replace("-", " ").title() if parts else ""
        data_path = REPO_ROOT / "app" / f"{route.strip('/')}.json"
        if not data_path.is_file() and route.endswith("/support-worker"):
            alternate = REPO_ROOT / "app" / f"{route.strip('/')}-jobs.json"
            data_path = alternate if alternate.is_file() else data_path
        add("Regional/category", area, family, route, _job_count(data_path) if data_path.is_file() else "")

    city_register = _load_json(PIPELINE_ROOT / "city_pages" / "city-page-register.json") or []
    for row in city_register:
        if row.get("lifecycle_state") != "active" or row.get("mode") != "publish":
            continue
        route = str(row.get("route") or "").strip()
        output = REPO_ROOT / str(row.get("output_json") or "")
        if route:
            add(
                "City",
                str(row.get("display_name") or ""),
                str(row.get("category_label") or ""),
                route,
                _job_count(output) if output.is_file() else 0,
            )

    broad_city_register = _load_json(BROAD_CITY_PAGE_REGISTER) or []
    for row in broad_city_register:
        if row.get("lifecycle_state") != "active":
            continue
        route = str(row.get("route") or "").strip()
        if route:
            add("City", str(row.get("display_name") or ""), "All roles", route)

    configured_files: list[Path] = []
    for (region, category), status in statuses.items():
        if status != "LIVE" or region not in regions or category not in categories:
            continue
        region_slug = regions[region]["slug"]
        category_slug = categories[category]["route_slug"]
        if (REPO_ROOT / "app" / region_slug / category_slug / "page.tsx").is_file():
            continue
        data_path = REPO_ROOT / "app" / "_city-pages" / "configured-slices" / region_slug / f"{category_slug}.json"
        jobs = _job_count(data_path) if data_path.is_file() else 0
        if not jobs:
            continue
        configured_files.append(data_path)
        add(
            "Regional/category",
            region,
            str(categories[category].get("display_label") or category),
            f"/job-search/{region_slug}/{category_slug}",
            jobs,
        )

    northern_ireland_headline = re.compile(
        r"\b(?:belfast|londonderry|derriaghy|northern ireland|l['’]?derry|derry)\b", re.I
    )
    northern_ireland_description = re.compile(
        r"\b(?:belfast|londonderry|derriaghy|l['’]?derry|derry)(?:\s+city\s+centre)?(?:-based|\s+based)?\b"
        r"|\bbased\s+(?:in|at)\s+(?:belfast|londonderry|derriaghy|l['’]?derry|derry)\b"
        r"|\bnorthern\s+ireland\b",
        re.I,
    )
    published_ids: set[str] = set()
    static_files = [
        path for path in (REPO_ROOT / "app").rglob("*.json")
        if "_city-pages" not in path.parts
    ]
    for path in sorted([*static_files, *configured_files]):
        data = _load_json(path)
        if not isinstance(data, list):
            continue
        for row in data:
            if not isinstance(row, dict):
                continue
            job_id = str(row.get("job_id") or "").strip()
            if not job_id or not str(row.get("title") or "").strip() or not str(row.get("apply_url") or "").strip():
                continue
            if str(row.get("region") or "").strip().lower() == "london":
                headline = f"{row.get('title', '')} {row.get('location', '')}"
                description = f"{row.get('description', '')} {row.get('full_description', '')}"
                if northern_ireland_headline.search(headline) or northern_ireland_description.search(description):
                    continue
            published_ids.add(job_id)

    counts = Counter(row[1] for row in details)
    counts["Individual job"] = len(published_ids)
    counts["Total"] = len(details) + len(published_ids)
    summary = [
        ["Summary", "All published/indexable URLs", "Sitewide", "", "", counts["Total"], "", "Yes"],
        ["Summary", "Individual job", "Sitewide", "", "/jobs/[id]", counts["Individual job"], counts["Individual job"], "Yes"],
        ["Summary", "Regional/category", "Sitewide", "", "Multiple", counts["Regional/category"], "", "Yes"],
        ["Summary", "City", "Sitewide", "", "Multiple", counts["City"], "", "Yes"],
        ["Summary", "Core", "Sitewide", "", "Multiple", counts["Core"], "", "Yes"],
    ]
    return [*summary, *details], dict(counts)


def _git_show_main(repo_path: str) -> str:
    try:
        return subprocess.check_output(["git", "show", f"origin/main:{repo_path}"], cwd=REPO_ROOT, text=True)
    except subprocess.CalledProcessError as exc:
        raise RuntimeError(f"Could not read {repo_path} from origin/main") from exc


def _main_tree_names(folder: str) -> list[str]:
    try:
        return subprocess.check_output(["git", "ls-tree", "-r", "--name-only", "origin/main", folder], cwd=REPO_ROOT, text=True).splitlines()
    except subprocess.CalledProcessError as exc:
        raise RuntimeError(f"Could not inspect origin/main/{folder}") from exc


def _latest_live_source_report() -> tuple[str, str]:
    candidates = sorted(
        name for name in _main_tree_names("pipeline/reports-daily")
        if name.startswith("pipeline/reports-daily/live-job-source-count-") and name.endswith(".csv")
    )
    if not candidates:
        raise RuntimeError("No live-job-source-count report found on origin/main")
    report_path = candidates[-1]
    return report_path, _git_show_main(report_path)


def _load_source_report_summary() -> SourceReportSummary:
    report_path, text = _latest_live_source_report()
    reader = csv.DictReader(io.StringIO(text))
    required = {"report_date", "level", "source", "count"}
    if not reader.fieldnames or not required.issubset(reader.fieldnames):
        raise RuntimeError(f"Unexpected live source report header in {report_path}")
    for row in reader:
        if (
            (row.get("level") or "").strip() == "total"
            and (row.get("source") or "").strip() == "All"
        ):
            return SourceReportSummary(
                path=report_path,
                report_date=(row.get("report_date") or "").strip(),
                total_live_jobs=int((row.get("count") or "0").strip()),
            )
    raise RuntimeError(f"No total/All row found in {report_path}")


def _site_inventory_summary(
    inventory: LiveInventory,
    *,
    report_date: str,
    rollups: dict[str, str],
    statuses: dict[tuple[str, str], str],
) -> SiteInventorySummary:
    jobs_by_id = {job.job_id: job for job in inventory.jobs}
    placements: set[tuple[str, str, str]] = set()

    for placement in inventory.placements:
        family_key = FAMILY_BY_FILENAME.get(Path(placement.source_file).name)
        if not family_key:
            continue
        region = rollups.get(placement.region, placement.region)
        placements.add((placement.canonical_job_id, region, family_key))

    placement_counts = Counter(job_id for job_id, _region, _family in placements)
    represented_jobs = set(placement_counts)
    all_jobs = set(jobs_by_id)
    outside = all_jobs - represented_jobs

    duplicate_jobs_by_provider: Counter[str] = Counter()
    extra_placements_by_provider: Counter[str] = Counter()
    for job_id, count in placement_counts.items():
        if count <= 1:
            continue
        provider = jobs_by_id[job_id].source
        duplicate_jobs_by_provider[provider] += 1
        extra_placements_by_provider[provider] += count - 1

    slice_counts: Counter[tuple[str, str]] = Counter()
    non_live_jobs: set[str] = set()
    family_register = {family["key"]: family["register_category"] for family in FAMILIES}
    for job_id, region, family_key in placements:
        register_category = family_register[family_key]
        if statuses.get((region, register_category), "") == "LIVE":
            slice_counts[(region, family_key)] += 1
        else:
            non_live_jobs.add(job_id)

    provider_counts = Counter(job.source for job in inventory.jobs)
    unique_jobs = len(inventory.jobs)
    unique_jobg8 = provider_counts.get("JobG8", 0)
    return SiteInventorySummary(
        report_date=report_date,
        unique_live_jobs=unique_jobs,
        unique_jobg8_jobs=unique_jobg8,
        unique_external_jobs=unique_jobs - unique_jobg8,
        provider_counts=dict(provider_counts),
        provider_duplicate_jobs=dict(duplicate_jobs_by_provider),
        provider_extra_placements=dict(extra_placements_by_provider),
        slice_counts=dict(slice_counts),
        slice_placements=len(placements),
        jobs_on_multiple_slices=sum(count > 1 for count in placement_counts.values()),
        extra_slice_placements=sum(count - 1 for count in placement_counts.values()),
        jobs_outside_governed_slices=len(outside),
        non_live_slice_jobs=len(non_live_jobs),
    )


def _live_count_for_market(
    live_counts: dict[tuple[str, str], int],
    rollups: dict[str, str],
    region_name: str,
    category: str,
) -> int:
    """Count canonical-market LIVE jobs, including exact geo aliases/detail regions.

    The live-source report records the factual region carried by each unique
    published vacancy. A canonical market may therefore have both direct rows
    (for example external sources already labelled ``North East``) and detail
    rows (for example JobG8 Tyneside or County Durham). Both belong to the same
    governed market and must contribute to its LIVE count.
    """
    total = live_counts.get((region_name, category), 0)
    total += sum(
        count
        for (raw_region, raw_category), count in live_counts.items()
        if raw_category == category and rollups.get(raw_region) == region_name
    )
    return total


def _load_selected_counts(repo_path: str) -> dict[str, int]:
    text = _git_show_main(repo_path)
    reader = csv.DictReader(io.StringIO(text.lstrip("\ufeff")))
    if not reader.fieldnames or "region" not in reader.fieldnames:
        raise RuntimeError(f"Unexpected decision report header in {repo_path}")
    selected_ids: dict[str, set[str]] = {}
    anonymous_counts: dict[str, int] = {}
    for row in reader:
        decision = (row.get("decision") or "").strip().upper()
        selection_status = (row.get("selection_status") or "").strip().upper()
        if decision != "SELECTED" and selection_status != "SELECTED":
            continue
        region = (row.get("region") or "").strip()
        if not region:
            continue
        job_id = (row.get("job_id") or "").strip()
        if job_id:
            selected_ids.setdefault(region, set()).add(job_id)
        else:
            anonymous_counts[region] = anonymous_counts.get(region, 0) + 1
    regions = set(selected_ids) | set(anonymous_counts)
    return {region: len(selected_ids.get(region, set())) + anonymous_counts.get(region, 0) for region in regions}


def _load_latest_profile_counts() -> tuple[str, str, dict[tuple[str, str], int]]:
    candidates = sorted(
        name for name in _main_tree_names("pipeline/reports-module2")
        if name.startswith("pipeline/reports-module2/") and name.endswith("-module2-daily-counts.csv")
    )
    if not candidates:
        return "", "", {}
    report_path = candidates[-1]
    text = _git_show_main(report_path)
    reader = csv.DictReader(io.StringIO(text.lstrip("\ufeff")))
    required = {"date", "region", "region_scope", "category", "daily_job_count"}
    if not reader.fieldnames or not required.issubset(reader.fieldnames):
        raise RuntimeError(f"Unexpected Module 2 daily-counts header in {report_path}")
    rows = list(reader)
    dates = sorted({(row.get("date") or "").strip() for row in rows if (row.get("date") or "").strip()})
    if not dates:
        return report_path, "", {}
    latest_date = dates[-1]
    counts: dict[tuple[str, str], int] = {}
    for row in rows:
        if (row.get("date") or "").strip() != latest_date or (row.get("region_scope") or "").strip() != "lookup_region":
            continue
        region = (row.get("region") or "").strip()
        category = (row.get("category") or "").strip()
        if region and category:
            counts[(region, category)] = int(float((row.get("daily_job_count") or "0").strip()))
    return report_path, latest_date, counts


def _load_teaching_vacancies_counts(regions: list[tuple[str, str]]) -> dict[str, int]:
    folder = "pipeline/output-external/teaching-vacancies-regional"
    available = set(_main_tree_names(folder))
    counts: dict[str, int] = {}
    for region_name, slug in regions:
        path = f"{folder}/{slug}-admin-service.json"
        if path not in available:
            continue
        data = json.loads(_git_show_main(path).lstrip("\ufeff"))
        counts[region_name] = _count_json_data(data)
    return counts


def build() -> str:
    catalog = _load_json(CATALOG)
    if not isinstance(catalog, dict) or not isinstance(catalog.get("regions"), dict):
        raise RuntimeError(f"Could not load assessable region catalogue: {CATALOG}")

    statuses = _load_statuses()
    source_report = _load_source_report_summary()
    jobg8_profile = _load_jobg8_category_profile()
    family_coverage_date = _load_family_coverage_date()
    if jobg8_profile and jobg8_profile.feed_date != family_coverage_date:
        raise RuntimeError(
            "Refusing mixed-date overview: JobG8 category profile is "
            f"{jobg8_profile.feed_date}, family coverage is {family_coverage_date}"
        )
    _profile_report, profile_date, profile_counts = _load_latest_profile_counts()
    decision_counts = {
        family["key"]: _load_selected_counts(family["decision_report"])
        for family in FAMILIES if family["decision_report"]
    }

    regions = sorted(((name, facts["slug"]) for name, facts in catalog["regions"].items()), key=lambda item: item[0].casefold())
    declared_count = int(catalog.get("region_count") or len(regions))
    if declared_count != len(regions) or len(regions) != EXPECTED_REGION_COUNT:
        raise RuntimeError(
            f"Daily overview expected {EXPECTED_REGION_COUNT} assessable UK markets, catalogue declares {declared_count} and contains {len(regions)}."
        )

    rollups = {str(k): str(v) for k, v in catalog.get("detail_rollups", {}).items()}
    london_now = datetime.now(ZoneInfo("Europe/London"))
    report_date = london_now.date().isoformat()
    inventory = collect_live_inventory(
        REPO_ROOT / "app",
        as_of=london_now.date(),
        now=london_now,
    )
    site = _site_inventory_summary(
        inventory,
        report_date=report_date,
        rollups=rollups,
        statuses=statuses,
    )
    page_rows, page_counts = _published_page_inventory()
    city_rows, city_mapped_jobs, city_unmapped_jobs = _city_opportunity_rows(inventory)
    teaching_counts = _load_teaching_vacancies_counts(regions)

    def profile_count(region_name: str, category: str) -> int | None:
        direct = profile_counts.get((region_name, category))
        if direct is not None:
            return direct
        alias_total = sum(
            count for (raw_region, raw_category), count in profile_counts.items()
            if raw_category == category and rollups.get(raw_region) == region_name
        )
        return alias_total if alias_total else None

    rows = []
    for region_name, slug in regions:
        family_state = {}
        for family in FAMILIES:
            status = statuses.get((region_name, family["register_category"]), "")
            is_live = status == "LIVE"
            live_count = site.slice_counts.get((region_name, family["key"]), 0) if is_live else 0

            candidate_count: int | None = None
            candidate_source = ""
            if not is_live:
                if family["decision_report"] and region_name in decision_counts[family["key"]]:
                    candidate_count = decision_counts[family["key"]][region_name]
                    candidate_source = "daily"
                elif family["profile_category"]:
                    candidate_count = profile_count(region_name, family["profile_category"])
                    if candidate_count is not None:
                        candidate_source = "profile"
                if family["key"] == "service_admin" and region_name in teaching_counts:
                    candidate_count = (candidate_count or 0) + teaching_counts[region_name]
                    candidate_source = (candidate_source + "+teaching").strip("+")

            family_state[family["key"]] = {
                "live": is_live,
                "live_count": live_count,
                "candidate_count": candidate_count,
                "candidate_source": candidate_source,
                "status": status,
            }
        rows.append((region_name, slug, family_state))

    source_report_status = (
        "CURRENT"
        if source_report.report_date == site.report_date
        and source_report.total_live_jobs == site.unique_live_jobs
        else (
            f"STALE — CSV says {source_report.total_live_jobs:,} for "
            f"{source_report.report_date or 'unknown date'}"
        )
    )
    provider_rows = []
    for provider, count in sorted(
        site.provider_counts.items(),
        key=lambda item: (item[0] != "JobG8", item[0].casefold()),
    ):
        provider_rows.append(
            f"| {provider} | {count:,} | "
            f"{site.provider_duplicate_jobs.get(provider, 0):,} | "
            f"{site.provider_extra_placements.get(provider, 0):,} |"
        )

    header = "| Region | " + " | ".join(family["label"] for family in FAMILIES) + " |"
    divider = "|---|" + "---:|" * len(FAMILIES)

    lines = [
        "# Ontap daily regional overview",
        "",
        f"Generated: {london_now.isoformat(timespec='seconds')}",
        "",
        "[Download this overview as Excel](./daily-region-overview.xlsx)",
        "",
        "## SITEWIDE RECONCILIATION",
        "",
        "| Measure | Count |",
        "|---|---:|",
        f"| Unique live jobs | {site.unique_live_jobs:,} |",
        f"| Unique JobG8 jobs | {site.unique_jobg8_jobs:,} |",
        f"| Unique non-JobG8 jobs | {site.unique_external_jobs:,} |",
        f"| Regional/category slice placements | {site.slice_placements:,} |",
        f"| Jobs appearing on multiple slices | {site.jobs_on_multiple_slices:,} |",
        f"| Extra slice placements | {site.extra_slice_placements:,} |",
        f"| Unique jobs outside governed slices | {site.jobs_outside_governed_slices:,} |",
        f"| Jobs found in non-LIVE slices | {site.non_live_slice_jobs:,} |",
        "",
        (
            f"**Reconciliation: {site.unique_live_jobs:,} unique jobs + "
            f"{site.extra_slice_placements:,} extra slice placements = "
            f"{site.slice_placements:,} regional/category slice placements.**"
            if site.jobs_outside_governed_slices == 0
            else (
                f"**CHECK: {site.jobs_outside_governed_slices:,} unique jobs are not represented "
                "on a governed regional/category slice.**"
            )
        ),
        "",
        f"Latest source-count CSV: `{source_report.path}` — **{source_report_status}**.",
        "",
        "### Provider breakdown",
        "",
        "| Provider | Unique live jobs | Jobs on 2+ slices | Extra slice placements |",
        "|---|---:|---:|---:|",
        *provider_rows,
        "",
    ]

    if jobg8_profile:
        lines.extend([
            "## JOBG8 FEED RECEIVED",
            "",
            f"**JobG8 jobs received: {jobg8_profile.total_jobs:,}** (feed date: {jobg8_profile.feed_date})",
            "",
            "| JobG8 classification | Jobs received | Ontap jobs |",
            "|---|---:|---:|",
            *(
                (
                    f"| {category} | {received:,} | {published:,} |"
                    if published else f"| {category} | {received:,} |  |"
                )
                for category, received, published in jobg8_profile.counts
            ),
            (
                f"| Total Ontap JobG8 jobs published today | {jobg8_profile.total_jobs:,} | "
                f"{jobg8_profile.published_jobs:,} |"
                if jobg8_profile.published_jobs is not None else
                f"| Total JobG8 jobs received | {jobg8_profile.total_jobs:,} |  |"
            ),
            "",
        ])

    lines.extend([
        "## PAGES",
        "",
        (
            f"**Published/indexable URLs: {page_counts['Total']:,}** — "
            f"{page_counts['Individual job']:,} individual job pages, "
            f"{page_counts['Regional/category']:,} regional/category pages, "
            f"{page_counts['City']:,} city pages and {page_counts['Core']:,} core pages."
        ),
        "",
        "| Level | Page type | Area | Family | URL | Page count | Live jobs | In sitemap |",
        "|---|---|---|---|---|---:|---:|---|",
        *(
            "| " + " | ".join(
                (f"{cell:,}" if isinstance(cell, int) else str(cell).replace("|", "\\|"))
                for cell in row
            ) + " |"
            for row in page_rows
        ),
        "",
    ])

    lines.extend([
        "## CITY OPPORTUNITIES",
        "",
        (
            f"**{len(city_rows):,} mapped towns/localities with live jobs or an existing city page.** "
            f"Counts use all {site.unique_live_jobs:,} unique live Ontap jobs across every role and provider: "
            f"{city_mapped_jobs:,} have an exact recognised town/locality and "
            f"{city_unmapped_jobs:,} have only broader or unrecognised location evidence. "
            f"CREATE means {CITY_PAGE_THRESHOLD}+ current jobs and no existing city page; London is held separately."
        ),
        "",
        "| Status | Town/city/locality | Region | All live jobs | Existing pages | Current routes |",
        "|---|---|---|---:|---:|---|",
        *(
            f"| {row.status} | {row.locality} | {row.region} | {row.live_jobs:,} | "
            f"{row.existing_pages:,} | {row.routes} |"
            for row in city_rows
        ),
        "",
    ])

    lines.extend([
        f"> LIVE counts come directly from the current published `app/` JSON, deduplicated within each canonical region/family slice while preserving legitimate appearances in more than one family. This is the live-site authority for the reconciliation above; the dated source-count CSV is shown only as a freshness cross-check. The overview covers all {EXPECTED_REGION_COUNT} assessable UK markets; LIVE status remains controlled only by the slice register. Before same-feed 78-market coverage has run, NOT LIVE Admin/Support and Customer Service may fall back to the latest all-region Module 2 profile ({profile_date or 'unavailable'}), and Service Admin may also add current Teaching Vacancies regional candidate output. `—` means not assessed / no current source; it does NOT mean zero.",
        "",
        "## LIVE",
        "",
        header,
        divider,
    ])

    for region_name, _slug, state in rows:
        live_cells = []
        for family in FAMILIES:
            item = state[family["key"]]
            if not item["live"]:
                live_cells.append("")
            elif item["live_count"] == 0:
                live_cells.append("CHECK")
            else:
                live_cells.append(str(item["live_count"]))
        lines.append(f"| {region_name} | " + " | ".join(live_cells) + " |")

    lines.extend(["", "## NOT LIVE", "", header, divider])
    for region_name, _slug, state in rows:
        cells = []
        for family in FAMILIES:
            item = state[family["key"]]
            if item["live"]:
                cells.append("")
            elif item["candidate_count"] is None:
                cells.append("—")
            else:
                cells.append(str(item["candidate_count"]))
        lines.append(f"| {region_name} | " + " | ".join(cells) + " |")

    live_regions = {f["key"]: sum(1 for _n, _s, state in rows if state[f["key"]]["live"]) for f in FAMILIES}
    live_jobs = {
        f["key"]: sum(state[f["key"]]["live_count"] for _n, _s, state in rows if state[f["key"]]["live"])
        for f in FAMILIES
    }
    checks = {
        f["key"]: sum(1 for _n, _s, state in rows if state[f["key"]]["live"] and state[f["key"]]["live_count"] == 0)
        for f in FAMILIES
    }
    total_live_slices = sum(live_regions.values())
    total_possible = len(rows) * len(FAMILIES)

    def headline_value(family_key: str) -> str:
        value = str(live_jobs[family_key])
        if checks[family_key]:
            value += f" + {checks[family_key]} CHECK"
        return value

    lines.extend([
        "", "## HEADLINE", "",
        "| Measure | " + " | ".join(family["label"] for family in FAMILIES) + " |",
        "|---|" + "---:|" * len(FAMILIES),
        "| Live regions | " + " | ".join(f"{live_regions[f['key']]} / {EXPECTED_REGION_COUNT}" for f in FAMILIES) + " |",
        "| Live slice placements | " + " | ".join(headline_value(f["key"]) for f in FAMILIES) + " |",
        "", f"**Live slices: {total_live_slices} / {total_possible}.**", "",
    ])
    return "\n".join(lines)


def main() -> None:
    OUTPUT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT.write_text(build(), encoding="utf-8")
    print(f"Wrote {OUTPUT}")


if __name__ == "__main__":
    main()

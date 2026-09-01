from __future__ import annotations

import argparse
from pathlib import Path
import re

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter


SECTION_SHEETS = (
    ("## SITEWIDE RECONCILIATION", "Sitewide"),
    ("## JOBG8 FEED RECEIVED", "JobG8 categories"),
    ("## LIVE", "LIVE"),
    ("## NOT LIVE", "NOT LIVE"),
)


def _cell_value(value: str):
    value = value.strip()
    if re.fullmatch(r"-?\d{1,3}(?:,\d{3})*", value):
        return int(value.replace(",", ""))
    return value


def _first_table_after(lines: list[str], heading: str) -> list[list[object]]:
    try:
        start = lines.index(heading) + 1
    except ValueError as exc:
        raise RuntimeError(f"Missing overview section: {heading}") from exc

    table_lines: list[str] = []
    for line in lines[start:]:
        if line.startswith("|"):
            table_lines.append(line)
        elif table_lines:
            break
    if len(table_lines) < 2:
        raise RuntimeError(f"No Markdown table found after: {heading}")

    rows = [
        [_cell_value(cell) for cell in line.strip().strip("|").split("|")]
        for line in table_lines
    ]
    return [rows[0], *rows[2:]]  # omit Markdown alignment row


def extract_tables(markdown: str) -> dict[str, list[list[object]]]:
    lines = markdown.splitlines()
    return {sheet: _first_table_after(lines, heading) for heading, sheet in SECTION_SHEETS}


def export(markdown_path: Path, output_path: Path) -> None:
    tables = extract_tables(markdown_path.read_text(encoding="utf-8"))
    workbook = Workbook()
    workbook.remove(workbook.active)

    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    alternating_fill = PatternFill("solid", fgColor="F3F6F9")
    light_border = Side(style="thin", color="D5DCE3")

    for sheet_name, rows in tables.items():
        sheet = workbook.create_sheet(sheet_name)
        sheet.sheet_view.showGridLines = False
        sheet.freeze_panes = "A2"
        for row in rows:
            sheet.append(row)

        for cell in sheet[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        sheet.row_dimensions[1].height = 32
        sheet.auto_filter.ref = sheet.dimensions

        for row_number in range(2, sheet.max_row + 1):
            if row_number % 2 == 0:
                for cell in sheet[row_number]:
                    cell.fill = alternating_fill
            for cell in sheet[row_number]:
                cell.border = Border(bottom=light_border)
                cell.alignment = Alignment(
                    horizontal="right" if isinstance(cell.value, int) else "left",
                    vertical="center" if sheet_name == "NOT LIVE" else "top",
                    wrap_text=sheet_name != "NOT LIVE",
                )
                if isinstance(cell.value, int):
                    cell.number_format = "#,##0"

        for column in range(1, sheet.max_column + 1):
            max_length = max(len(str(sheet.cell(row, column).value or "")) for row in range(1, sheet.max_row + 1))
            sheet.column_dimensions[get_column_letter(column)].width = min(max(max_length + 3, 12), 34)
        if sheet_name in {"LIVE", "NOT LIVE"}:
            sheet.column_dimensions["A"].width = 34
            for column in range(2, sheet.max_column + 1):
                sheet.column_dimensions[get_column_letter(column)].width = 24 if sheet_name == "NOT LIVE" else 18
        if sheet_name == "NOT LIVE":
            for row_number in range(2, sheet.max_row + 1):
                sheet.row_dimensions[row_number].height = 20
        if sheet_name == "JobG8 categories" and sheet.max_column >= 3:
            green_fill = PatternFill("solid", fgColor="E2F0D9")
            for row_number in range(1, sheet.max_row + 1):
                sheet.cell(row_number, 3).fill = green_fill
            sheet.cell(1, 3).font = Font(color="000000", bold=False)
            sheet.cell(1, 3).alignment = Alignment(horizontal="left", vertical="center")
            for cell in sheet[sheet.max_row]:
                cell.fill = green_fill
                cell.font = Font(bold=True)

    output_path.parent.mkdir(parents=True, exist_ok=True)
    workbook.save(output_path)

    check = load_workbook(output_path, read_only=True, data_only=True)
    if check.sheetnames != [sheet for _heading, sheet in SECTION_SHEETS]:
        raise RuntimeError("Exported overview workbook has unexpected sheets")
    check.close()


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--markdown", required=True, type=Path)
    parser.add_argument("--output", required=True, type=Path)
    args = parser.parse_args()
    export(args.markdown, args.output)
    print(f"Wrote {args.output}")


if __name__ == "__main__":
    main()

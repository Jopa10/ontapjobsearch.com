#!/usr/bin/env python3
"""Export every row in the latest JobG8 feed with Ontap selection evidence.

This is an owner-facing diagnostic only. It does not change registers, selection,
published JSON or LIVE state.
"""

from __future__ import annotations

import argparse
import csv
import json
import re
from collections import Counter
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.formatting.rule import FormulaRule
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

from pipeline.scripts.jobg8_refine_other_families import annualised_salary, latest_feed, norm, text
from pipeline.scripts import service_admin_pipeline_core as selector_salary
from pipeline.scripts.pipeline_refinement import ANNUAL_SALARY_FACTORS


COL = {
    "id": "/Job/DisplayReference",
    "title": "/Job/Position",
    "description": "/Job/Description",
    "employer": "/Job/AdvertiserName",
    "advertiser_type": "/Job/AdvertiserType",
    "location": "/Job/Location",
    "area": "/Job/Area",
    "category": "/Job/Classification",
    "salary_min": "/Job/SalaryMinimum",
    "salary_max": "/Job/SalaryMaximum",
    "salary_period": "/Job/SalaryPeriod",
    "employment_type": "/Job/EmploymentType",
    "work_hours": "/Job/WorkHours",
    "apply_url": "/Job/ApplicationURL",
    "sender_reference": "/Job/SenderReference",
}

CATEGORY_LABELS = {
    "admin_service": "Admin / Customer Service",
    "support_worker": "Support Worker",
    "customer_service_contact_centre": "Customer Service / Contact Centre",
    "finance_accounts": "Finance / Accounts",
    "hr_recruitment": "HR / Recruitment",
    "warehouse_logistics": "Warehouse / Logistics",
    "marketing": "Marketing",
    "customer_sales": "Customer Sales / Sales Advisor",
    "legal_assistant_paralegal": "Legal Assistant / Paralegal",
}

OUTPUT_COLUMNS = [
    "JobG8 ID", "Title", "Employer", "Advertiser type", "Location", "JobG8 area",
    "Ontap region", "Original JobG8 category", "Employment type", "Work hours",
    "Salary minimum", "Salary maximum", "Salary period", "Annualised minimum",
    "Salary source", "Effective salary", "Annualised maximum", "Annualised midpoint",
    "Salary band", "Refined broad family",
    "Governed family matches", "Publication / coverage status", "Coverage evidence", "Application URL",
    "Sender reference", "Full description",
]


def split_values(value: object) -> list[str]:
    return [part.strip() for part in text(value).split(";") if part.strip()]


def conflict_categories(conflicts: list[str]) -> set[str]:
    return {item.split(":", 1)[0].strip() for item in conflicts if ":" in item}


def salary_band(minimum: float | None, maximum: float | None) -> str:
    values = [value for value in (minimum, maximum) if value is not None]
    if not values:
        return "Below £20k / unknown"
    midpoint = sum(values) / len(values)
    if midpoint < 20_000:
        return "Below £20k / unknown"
    if midpoint < 35_000:
        return "£20k–<£35k"
    if midpoint <= 45_000:
        return "£35k–£45k"
    return "Over £45k"


def effective_salary(row: pd.Series) -> tuple[str, str, float | None, float | None]:
    """Return the same structured/description/missing salary evidence used by selectors."""
    salary_text, salary_source = selector_salary.build_salary_details(row)
    if salary_source == "structured":
        period = selector_salary.normalise_salary_period(row)
        annual_min = annualised_salary(row.get(COL["salary_min"], ""), period)
        annual_max = annualised_salary(row.get(COL["salary_max"], ""), period)
        if annual_min is not None or annual_max is not None:
            return salary_source, salary_text, annual_min, annual_max
        # SalaryAdditional is structured evidence too.  When JobG8 leaves its
        # numeric min/max blank, annualise the effective text just as the live
        # selector does instead of incorrectly leaving the audit band unknown.
    if salary_source != "description_fallback":
        if not salary_text:
            return salary_source, salary_text, None, None

    amounts = [
        float(value.replace(",", ""))
        for value in re.findall(r"£\s*(\d[\d,]*(?:\.\d+)?)", salary_text)
    ]
    period = next(
        (
            candidate
            for candidate in ANNUAL_SALARY_FACTORS
            if re.search(rf"\b{re.escape(candidate)}\b", salary_text, flags=re.IGNORECASE)
        ),
        "",
    )
    if not amounts or not period:
        return salary_source, salary_text, None, None
    annualised = [amount * ANNUAL_SALARY_FACTORS[period] for amount in amounts]
    return salary_source, salary_text, min(annualised), max(annualised)


def published_jobg8_ids(app_root: Path) -> set[str]:
    ids: set[str] = set()
    for path in app_root.rglob("*.json"):
        try:
            payload = json.loads(path.read_text(encoding="utf-8-sig"))
        except (OSError, json.JSONDecodeError):
            continue
        rows: list[Any] = payload if isinstance(payload, list) else []
        if isinstance(payload, dict):
            rows = next((payload[key] for key in ("jobs", "items", "results") if isinstance(payload.get(key), list)), [])
        for row in rows:
            if isinstance(row, dict) and norm(row.get("source")) == "jobg8":
                job_id = text(row.get("job_id") or row.get("id"))
                if job_id:
                    ids.add(job_id)
    return ids


def published_jobg8_details(app_root: Path) -> dict[str, dict[str, str]]:
    details: dict[str, dict[str, str]] = {}
    for path in app_root.rglob("*.json"):
        try:
            payload = json.loads(path.read_text(encoding="utf-8-sig"))
        except (OSError, json.JSONDecodeError):
            continue
        for row in payload if isinstance(payload, list) else []:
            if not isinstance(row, dict) or norm(row.get("source")) != "jobg8":
                continue
            job_id = text(row.get("job_id") or row.get("id"))
            if job_id and job_id not in details:
                details[job_id] = {
                    "JobG8 ID": job_id, "Title": text(row.get("title")),
                    "Employer": text(row.get("company") or row.get("advertiser_name")),
                    "Ontap posted date": text(row.get("posted_date")),
                    "Apply URL": text(row.get("apply_url")),
                }
    return details


def write_published_absent_current_feed(raw: pd.DataFrame, input_dir: Path, published: dict[str, dict[str, str]], path: Path) -> None:
    current_ids = {text(value) for value in raw[COL["id"]] if text(value)}
    absent_ids = set(published) - current_ids
    last_seen: dict[str, str] = {}
    for feed in sorted(input_dir.glob("*.xlsx")):
        match = re.search(r"(20\d{2}-\d{2}-\d{2})", feed.stem)
        feed_date = match.group(1) if match else feed.stem
        for value in pd.read_excel(feed, usecols=[COL["id"]], dtype=str).fillna("")[COL["id"]]:
            job_id = text(value)
            if job_id in absent_ids:
                last_seen[job_id] = feed_date
    fields = ["JobG8 ID", "Title", "Employer", "Ontap posted date", "Last seen in JobG8 archive", "Apply URL"]
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=fields)
        writer.writeheader()
        for job_id in sorted(absent_ids, key=lambda value: (last_seen.get(value, ""), value), reverse=True):
            row = dict(published[job_id])
            row["Last seen in JobG8 archive"] = last_seen.get(job_id, "Not found in available archive")
            writer.writerow(row)


def load_register(path: Path) -> dict[tuple[str, str], str]:
    out: dict[tuple[str, str], str] = {}
    with path.open(encoding="utf-8-sig", newline="") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames != ["region", "category", "status"]:
            raise SystemExit(f"Unexpected slice register header: {reader.fieldnames}")
        for row in reader:
            out[(text(row["region"]), text(row["category"]))] = text(row["status"]).upper()
    return out


def load_geo(path: Path) -> tuple[dict[str, str], dict[str, str]]:
    areas = pd.read_excel(path, dtype=str).fillna("")
    area_lookup = {norm(r["Area"]): text(r["Cluster"]) for _, r in areas.iterrows() if norm(r.get("Area"))}
    fallbacks = pd.read_excel(path, sheet_name="LocationFallback", dtype=str).fillna("")
    location_lookup = {
        norm(r["Location"]): text(r["Cluster"])
        for _, r in fallbacks.iterrows()
        if norm(r.get("Status")) == "auto" and norm(r.get("Location"))
    }
    return area_lookup, location_lookup


def resolve_region(row: pd.Series, areas: dict[str, str], locations: dict[str, str]) -> str:
    area = norm(row.get(COL["area"], ""))
    if area and area not in {"not specified", "unknown", "city"}:
        return areas.get(area, "Other / Unknown")
    return locations.get(norm(row.get(COL["location"], "")), "Other / Unknown")


def classify_status(
    job_id: str,
    region: str,
    selected: list[str],
    conflicts: list[str],
    published_ids: set[str],
    register: dict[tuple[str, str], str],
) -> tuple[str, str]:
    if job_id in published_ids:
        return "Published", "JobG8 ID is present in the current published app JSON."
    if selected:
        live = [category for category in selected if register.get((region, category)) == "LIVE"]
        labels = ", ".join(CATEGORY_LABELS.get(x, x) for x in selected)
        if not live:
            return "Governed match; market not LIVE", f"Matched governed family register(s): {labels}; none is LIVE in {region}. This is coverage evidence, not a production-selector decision."
        live_labels = ", ".join(CATEGORY_LABELS.get(x, x) for x in live)
        return (
            "Governed match in LIVE market; not published",
            f"Matched a governed family that is LIVE in {region}: {live_labels}; JobG8 ID is absent from current published JSON. The reason is not established because this audit does not execute the production family selector.",
        )
    if conflicts:
        return "Governed register rejection", "; ".join(conflicts)
    return "No governed register match", "No governed-family register match was found in the coverage audit. This is not a production-selector decision."


def build_rows(
    raw: pd.DataFrame,
    reconciliation: pd.DataFrame,
    published_ids: set[str],
    register: dict[tuple[str, str], str],
    areas: dict[str, str],
    locations: dict[str, str],
) -> list[dict[str, Any]]:
    by_title = {norm(row["title"]): row for _, row in reconciliation.iterrows()}
    rows: list[dict[str, Any]] = []
    for _, source in raw.iterrows():
        title = text(source.get(COL["title"], ""))
        evidence = by_title.get(norm(title), {})
        conflicts = split_values(evidence.get("refinement_conflicts", ""))
        selected = [
            category for category in split_values(evidence.get("selected_categories", ""))
            if category not in conflict_categories(conflicts)
        ]
        region = resolve_region(source, areas, locations)
        job_id = text(source.get(COL["id"], ""))
        status, reason = classify_status(job_id, region, selected, conflicts, published_ids, register)
        salary_source, effective_salary_text, annual_min, annual_max = effective_salary(source)
        annual_values = [v for v in (annual_min, annual_max) if v is not None]
        rows.append({
            "JobG8 ID": job_id,
            "Title": title,
            "Employer": text(source.get(COL["employer"], "")),
            "Advertiser type": text(source.get(COL["advertiser_type"], "")),
            "Location": text(source.get(COL["location"], "")),
            "JobG8 area": text(source.get(COL["area"], "")),
            "Ontap region": region,
            "Original JobG8 category": text(source.get(COL["category"], "")),
            "Employment type": text(source.get(COL["employment_type"], "")),
            "Work hours": text(source.get(COL["work_hours"], "")),
            "Salary minimum": text(source.get(COL["salary_min"], "")),
            "Salary maximum": text(source.get(COL["salary_max"], "")),
            "Salary period": text(source.get(COL["salary_period"], "")),
            "Annualised minimum": annual_min,
            "Salary source": salary_source,
            "Effective salary": effective_salary_text,
            "Annualised maximum": annual_max,
            "Annualised midpoint": (sum(annual_values) / len(annual_values)) if annual_values else None,
            "Salary band": salary_band(annual_min, annual_max),
            "Refined broad family": text(evidence.get("refined_broad_family", evidence.get("primary_broad_family", "Other / Unclassified"))),
            "Governed family matches": "; ".join(CATEGORY_LABELS.get(x, x) for x in selected),
            "Publication / coverage status": status,
            "Coverage evidence": reason,
            "Application URL": text(source.get(COL["apply_url"], "")),
            "Sender reference": text(source.get(COL["sender_reference"], "")),
            "Full description": text(source.get(COL["description"], "")),
        })
    return rows


def write_csv(rows: list[dict[str, Any]], path: Path) -> None:
    with path.open("w", encoding="utf-8-sig", newline="") as handle:
        writer = csv.DictWriter(handle, fieldnames=OUTPUT_COLUMNS)
        writer.writeheader()
        writer.writerows(rows)


def write_jobg8_category_profile(
    rows: list[dict[str, Any]],
    path: Path,
    feed_date: str,
    published_ids: set[str] | None = None,
) -> None:
    """Persist supplied and currently published counts by JobG8 classification."""
    counts = Counter((row.get("Original JobG8 category") or "(blank)").strip() for row in rows)
    published_counts: Counter[str] = Counter()
    matched_published_ids: set[str] = set()
    for row in rows:
        if row.get("Publication / coverage status") != "Published":
            continue
        job_id = text(row.get("JobG8 ID"))
        if job_id and job_id in matched_published_ids:
            continue
        if job_id:
            matched_published_ids.add(job_id)
        published_counts[(row.get("Original JobG8 category") or "(blank)").strip()] += 1
    if published_ids is not None:
        missing_ids = published_ids - matched_published_ids
        if missing_ids:
            published_counts["Published JobG8 ID absent from current feed"] = len(missing_ids)
        published_total = len(published_ids)
    else:
        published_total = sum(published_counts.values())
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8", newline="") as handle:
        writer = csv.DictWriter(
            handle,
            fieldnames=[
                "feed_date", "total_jobs", "published_jobg8_jobs",
                "jobg8_category", "count", "published_count",
            ],
        )
        writer.writeheader()
        categories = set(counts) | set(published_counts)
        for category in sorted(categories, key=lambda item: (-counts.get(item, 0), item.casefold())):
            writer.writerow({
                "feed_date": feed_date,
                "total_jobs": len(rows),
                "published_jobg8_jobs": published_total,
                "jobg8_category": category,
                "count": counts.get(category, 0),
                "published_count": published_counts.get(category, 0),
            })


def write_xlsx(rows: list[dict[str, Any]], path: Path, source_name: str) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "JobG8 selection audit"
    ws.sheet_view.showGridLines = False
    counts = Counter(row["Publication / coverage status"] for row in rows)
    ws.append(["JobG8 selection audit", "", "Source", source_name, "Rows", len(rows)])
    ws.merge_cells("A1:B1")
    ws.append(["Published", counts["Published"], "Not published", len(rows) - counts["Published"]])
    ws.append(["Diagnostic only", "No selection, register, LIVE state or published JSON is changed."])
    ws.append([])
    ws.append(OUTPUT_COLUMNS)
    for row in rows:
        ws.append([row[column] for column in OUTPUT_COLUMNS])

    ws.freeze_panes = "C6"
    ws.auto_filter.ref = f"A5:{get_column_letter(len(OUTPUT_COLUMNS))}{len(rows)+5}"
    table = Table(displayName="JobG8SelectionAudit", ref=f"A5:{get_column_letter(len(OUTPUT_COLUMNS))}{len(rows)+5}")
    table.tableStyleInfo = TableStyleInfo(name="TableStyleMedium2", showRowStripes=True, showFirstColumn=False, showLastColumn=False)
    ws.add_table(table)
    ws.row_dimensions[1].height = 28
    ws["A1"].font = Font(size=16, bold=True, color="FFFFFF")
    for cell in ws[1]:
        cell.fill = PatternFill("solid", fgColor="17324D")
        cell.font = Font(color="FFFFFF", bold=True)
    for cell in ws[5]:
        cell.alignment = Alignment(wrap_text=True, vertical="center")
    widths = [18, 30, 24, 16, 18, 24, 25, 28, 18, 16, 15, 15, 15, 18, 18, 30, 18, 18, 22, 30, 30, 42, 58, 38, 18, 90]
    for index, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(index)].width = width
    for row_number in range(6, len(rows) + 6):
        ws.row_dimensions[row_number].height = 48
        for col in (2, 3, 5, 6, 7, 8, 16, 20, 21, 22, 23, 26):
            ws.cell(row_number, col).alignment = Alignment(vertical="top", wrap_text=True)
        for col in (14, 17, 18):
            ws.cell(row_number, col).number_format = '£#,##0'
    status_col = get_column_letter(OUTPUT_COLUMNS.index("Publication / coverage status") + 1)
    data_range = f"{status_col}6:{status_col}{len(rows)+5}"
    for phrase, colour in (("Published", "DDF3E4"), ("market not LIVE", "FFF1C7"), ("not published", "FFE7C2"), ("rejection", "FDE2E2"), ("No governed", "E7E9ED")):
        ws.conditional_formatting.add(data_range, FormulaRule(formula=[f'ISNUMBER(SEARCH("{phrase}",{status_col}6))'], fill=PatternFill("solid", fgColor=colour)))
    wb.save(path)


def main() -> int:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input-dir", required=True, type=Path)
    parser.add_argument("--reconciliation-csv", required=True, type=Path)
    parser.add_argument("--slice-register", type=Path, default=Path("pipeline/registers/region_category_slice_register.csv"))
    parser.add_argument("--geo-lookup", type=Path, default=Path("pipeline/geo/geo_lookup.xlsx"))
    parser.add_argument("--app-root", type=Path, default=Path("app"))
    parser.add_argument("--output-dir", required=True, type=Path)
    parser.add_argument("--category-profile-output", type=Path)
    parser.add_argument("--limit", type=int)
    args = parser.parse_args()

    feed = latest_feed(args.input_dir)
    raw = pd.read_excel(feed, dtype=str).fillna("")
    if args.limit:
        raw = raw.head(args.limit)
    missing = [column for column in COL.values() if column not in raw.columns]
    if missing:
        raise SystemExit(f"JobG8 feed is missing expected columns: {missing}")
    reconciliation = pd.read_csv(args.reconciliation_csv, dtype=str, encoding="utf-8-sig").fillna("")
    areas, locations = load_geo(args.geo_lookup)
    published_details = published_jobg8_details(args.app_root)
    published_ids = set(published_details)
    rows = build_rows(raw, reconciliation, published_ids, load_register(args.slice_register), areas, locations)

    args.output_dir.mkdir(parents=True, exist_ok=True)
    csv_path = args.output_dir / "jobg8-selection-audit.csv"
    xlsx_path = args.output_dir / "jobg8-selection-audit.xlsx"
    write_csv(rows, csv_path)
    write_xlsx(rows, xlsx_path, feed.name)
    write_published_absent_current_feed(raw, args.input_dir, published_details, args.output_dir / "jobg8-published-absent-current-feed.csv")
    if args.category_profile_output:
        feed_date_match = re.search(r"(20\d{2}-\d{2}-\d{2})", feed.stem)
        feed_date = feed_date_match.group(1) if feed_date_match else feed.stem
        write_jobg8_category_profile(rows, args.category_profile_output, feed_date, published_ids)
    counts = Counter(row["Publication / coverage status"] for row in rows)
    if sum(counts.values()) != len(raw) or len(rows) != len(raw):
        raise SystemExit("Audit row reconciliation failed")
    print(f"Exported {len(rows):,} JobG8 rows from {feed.name}")
    for status, count in counts.most_common():
        print(f"  {status}: {count:,}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

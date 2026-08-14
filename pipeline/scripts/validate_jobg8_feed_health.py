from __future__ import annotations

import argparse
from dataclasses import dataclass
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import load_workbook


DISPLAY_REFERENCE = "/Job/DisplayReference"
POSITION = "/Job/Position"
ADVERTISER_NAME = "/Job/AdvertiserName"
AREA = "/Job/Area"
LOCATION = "/Job/Location"
APPLICATION_URL = "/Job/ApplicationURL"
DESCRIPTION = "/Job/Description"

REQUIRED_COLUMNS = {
    DISPLAY_REFERENCE,
    POSITION,
    ADVERTISER_NAME,
    AREA,
    LOCATION,
    APPLICATION_URL,
    DESCRIPTION,
}


@dataclass(frozen=True)
class FeedHealth:
    rows: int
    reference_ratio: float
    title_ratio: float
    advertiser_ratio: float
    geo_ratio: float
    application_url_ratio: float
    valid_http_url_ratio: float
    description_ratio: float
    unique_reference_ratio: float


def present(value: object) -> bool:
    return value is not None and str(value).strip() != ""


def valid_http_url(value: object) -> bool:
    if not present(value):
        return False
    try:
        parsed = urlparse(str(value).strip())
    except ValueError:
        return False
    return parsed.scheme.lower() in {"http", "https"} and bool(parsed.netloc)


def ratio(count: int, total: int) -> float:
    return count / total if total else 0.0


def inspect_feed(path: Path) -> FeedHealth:
    if not path.is_file() or path.stat().st_size <= 0:
        raise RuntimeError(f"JobG8 feed workbook is missing or empty: {path}")

    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        sheet = workbook.active
        row_iter = sheet.iter_rows(values_only=True)
        try:
            headers = next(row_iter)
        except StopIteration as exc:
            raise RuntimeError("JobG8 feed workbook contains no header row") from exc

        header_map = {
            str(value).strip(): index
            for index, value in enumerate(headers)
            if value is not None and str(value).strip()
        }
        missing = sorted(REQUIRED_COLUMNS.difference(header_map))
        if missing:
            raise RuntimeError("Missing JobG8 health-check fields: " + ", ".join(missing))

        rows = 0
        references = 0
        titles = 0
        advertisers = 0
        geos = 0
        application_urls = 0
        valid_urls = 0
        descriptions = 0
        reference_values: list[str] = []

        for values in row_iter:
            rows += 1

            reference = values[header_map[DISPLAY_REFERENCE]]
            title = values[header_map[POSITION]]
            advertiser = values[header_map[ADVERTISER_NAME]]
            area = values[header_map[AREA]]
            location = values[header_map[LOCATION]]
            app_url = values[header_map[APPLICATION_URL]]
            description = values[header_map[DESCRIPTION]]

            if present(reference):
                references += 1
                reference_values.append(str(reference).strip())
            if present(title):
                titles += 1
            if present(advertiser):
                advertisers += 1
            if present(area) or present(location):
                geos += 1
            if present(app_url):
                application_urls += 1
                if valid_http_url(app_url):
                    valid_urls += 1
            if present(description):
                descriptions += 1

        unique_references = len(set(reference_values))
        return FeedHealth(
            rows=rows,
            reference_ratio=ratio(references, rows),
            title_ratio=ratio(titles, rows),
            advertiser_ratio=ratio(advertisers, rows),
            geo_ratio=ratio(geos, rows),
            application_url_ratio=ratio(application_urls, rows),
            valid_http_url_ratio=ratio(valid_urls, application_urls),
            description_ratio=ratio(descriptions, rows),
            unique_reference_ratio=ratio(unique_references, references),
        )
    finally:
        workbook.close()


def validate_health(health: FeedHealth) -> list[str]:
    # Deliberately conservative. These thresholds are intended to catch a
    # structurally damaged feed, not normal commercial/content fluctuations.
    checks = [
        ("non-blank DisplayReference", health.reference_ratio, 0.95),
        ("non-blank Position/title", health.title_ratio, 0.95),
        ("non-blank ApplicationURL", health.application_url_ratio, 0.90),
        ("valid http(s) ApplicationURL among populated URLs", health.valid_http_url_ratio, 0.90),
        ("non-blank Description", health.description_ratio, 0.75),
        ("non-blank AdvertiserName", health.advertiser_ratio, 0.75),
        ("Area or Location populated", health.geo_ratio, 0.75),
        ("unique DisplayReference among populated references", health.unique_reference_ratio, 0.90),
    ]
    failures: list[str] = []
    for label, actual, minimum in checks:
        if actual < minimum:
            failures.append(f"{label}: {actual:.1%} < required {minimum:.1%}")
    return failures


def print_health(health: FeedHealth) -> None:
    print(f"JobG8 health rows: {health.rows}")
    print(f"DisplayReference populated: {health.reference_ratio:.1%}")
    print(f"Position/title populated: {health.title_ratio:.1%}")
    print(f"AdvertiserName populated: {health.advertiser_ratio:.1%}")
    print(f"Area or Location populated: {health.geo_ratio:.1%}")
    print(f"ApplicationURL populated: {health.application_url_ratio:.1%}")
    print(f"Valid http(s) URLs among populated ApplicationURL: {health.valid_http_url_ratio:.1%}")
    print(f"Description populated: {health.description_ratio:.1%}")
    print(f"Unique DisplayReference among populated references: {health.unique_reference_ratio:.1%}")


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Fail safely when a JobG8 feed has valid shape/count but catastrophic field quality."
    )
    parser.add_argument("--input", required=True, type=Path)
    args = parser.parse_args()

    health = inspect_feed(args.input)
    print_health(health)
    failures = validate_health(health)
    if failures:
        print("JobG8 feed health guard FAILED:")
        for failure in failures:
            print(f"- {failure}")
        raise SystemExit(1)

    print("JobG8 feed health guard: PASS")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

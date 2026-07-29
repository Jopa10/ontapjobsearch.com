from datetime import date, datetime
import inspect
import json
from pathlib import Path
import sys
import tempfile
import unittest
from urllib.parse import parse_qs, urlsplit

from openpyxl import Workbook


PIPELINE_ROOT = Path(__file__).resolve().parents[1]
if str(PIPELINE_ROOT) not in sys.path:
    sys.path.insert(0, str(PIPELINE_ROOT))

from external_sources.compose_northeast_admin import compose_rows  # noqa: E402
from external_sources.northeast_jobs_poc import (  # noqa: E402
    ManualDecisionState,
    Vacancy,
    approval_errors,
    approved_output_rows,
    main as nejobs_main,
    review_fingerprint,
    vacancy_to_published_job,
)
from external_sources import northeast_jobs_poc  # noqa: E402


def vacancy(
    source_job_id: str,
    *,
    closing_date: str = "31/08/2026 23:59",
    classification: str = "HC",
) -> Vacancy:
    return Vacancy(
        source="North East Jobs",
        source_job_id=source_job_id,
        title="Administration Assistant",
        employer="Example Council",
        location="Durham (derived for filtering)",
        ontap_geography="North East - County Durham & Darlington/Hartlepool",
        contract_type="Permanent",
        working_pattern="Full time",
        salary_text="£25,000 - £27,000",
        posted_date="29/07/2026",
        closing_date=closing_date,
        source_url=f"https://www.northeastjobs.org.uk/job/Test/{source_job_id}",
        screening_basis="clear title",
        detail_status="snapshot",
        classification=classification,
        classification_reason="clear transferable title",
        duplicate_status="UNIQUE",
    )


def decisions(*reviewed_ids: str) -> ManualDecisionState:
    return ManualDecisionState(
        selections=set(),
        exclusions=set(),
        reviewed_ids=set(reviewed_ids),
        review_date="2026-07-29",
    )


class NEJobsApprovedPublishTests(unittest.TestCase):
    def test_review_then_approved_output_end_to_end(self):
        if "sheet_name" not in inspect.signature(
            northeast_jobs_poc.pd.read_excel
        ).parameters:
            self.skipTest("another test replaced pandas with a minimal stub")
        with tempfile.TemporaryDirectory() as temporary:
            root = Path(temporary)
            rss = root / "rss.xml"
            details = root / "details"
            details.mkdir()
            jobg8 = root / "jobg8.xlsx"
            geo = root / "geo.xlsx"
            review_csv = root / "review.csv"
            review_md = root / "review.md"
            approved_json = root / "approved.json"

            rss.write_text(
                """<?xml version="1.0"?>
                <rss><channel><item>
                  <title>Administration Assistant</title>
                  <link>https://www.northeastjobs.org.uk/job/Test/300001</link>
                  <description><![CDATA[
                    Office administration
                    Contract Type: Permanent|Working Pattern: Full time|
                    Salary: £25,000|Advert End Date: 31/12/2099 23:59|
                  ]]></description>
                  <pubDate>Wed, 29 Jul 2026 00:00:00 GMT</pubDate>
                </item></channel></rss>""",
                encoding="utf-8",
            )
            (details / "300001.txt").write_text(
                """Title: Administration Assistant - Example Council

Contract Type:

Permanent

Working Pattern:

Full time

Advert Start Date:

29/07/2026

Advert End Date:

31/12/2099 23:59

Salary:

£25,000

Vacancy ID:

300001

Employment Location:

Durham
""",
                encoding="utf-8",
            )
            jobg8_book = Workbook()
            jobg8_sheet = jobg8_book.active
            jobg8_sheet.append(
                [
                    "/Job/DisplayReference",
                    "/Job/Position",
                    "/Job/AdvertiserName",
                    "/Job/Area",
                    "/Job/Location",
                    "/Job/Description",
                ]
            )
            jobg8_book.save(jobg8)

            geo_book = Workbook()
            area_sheet = geo_book.active
            area_sheet.title = "Sheet1"
            area_sheet.append(["Area", "Cluster"])
            area_sheet.append(
                [
                    "Durham",
                    "North East - County Durham & Darlington/Hartlepool",
                ]
            )
            fallback_sheet = geo_book.create_sheet("LocationFallback")
            fallback_sheet.append(["Location", "Cluster", "Status"])
            geo_book.save(geo)

            common = [
                "--jobg8",
                str(jobg8),
                "--geo-lookup",
                str(geo),
                "--rss-file",
                str(rss),
                "--details-dir",
                str(details),
                "--report-csv",
                str(review_csv),
                "--summary-md",
                str(review_md),
            ]
            self.assertEqual(0, nejobs_main(common))
            self.assertFalse(approved_json.exists())

            self.assertEqual(
                0,
                nejobs_main(
                    common
                    + [
                        "--write-approved-json",
                        "--approved-json",
                        str(approved_json),
                        "--confirm-approved",
                        "PUBLISH",
                    ]
                ),
            )
            rows = json.loads(approved_json.read_text(encoding="utf-8"))
            self.assertEqual(["nejobs-300001"], [row["job_id"] for row in rows])

    def test_approval_requires_exact_reviewed_set_and_no_fetch_failures(self):
        current = [vacancy("300001"), vacancy("300002", classification="POSS")]
        exact_decisions = decisions("300001", "300002")
        exact_decisions.review_fingerprint = review_fingerprint(current)

        self.assertEqual(
            [],
            approval_errors(
                current,
                exact_decisions,
                review_date="2026-07-29",
                failures=[],
            ),
        )

        incomplete_decisions = decisions("300001")
        incomplete_decisions.review_fingerprint = review_fingerprint(current)
        errors = approval_errors(
            current,
            incomplete_decisions,
            review_date="2026-07-29",
            failures=["300002 failed"],
        )
        self.assertTrue(any("new IDs: 300002" in error for error in errors))
        self.assertTrue(any("detail page(s) failed" in error for error in errors))

    def test_published_record_uses_facts_original_text_tracking_and_stable_id(self):
        row = vacancy_to_published_job(vacancy("300001"))
        query = parse_qs(urlsplit(row["apply_url"]).query)

        self.assertEqual("nejobs-300001", row["job_id"])
        self.assertEqual("NEJobs", row["source"])
        self.assertEqual("Durham", row["location"])
        self.assertEqual("2026-07-29", row["posted_date"])
        self.assertEqual("2026-08-31", row["closing_date"])
        self.assertTrue(
            row["summary"].startswith(
                "Administration Assistant with Example Council in Durham."
            )
        )
        self.assertGreaterEqual(len(row["description"]), 200)
        self.assertIn("complete duties, person specification", row["description"])
        self.assertEqual(["ontap"], query["utm_source"])
        self.assertEqual(["referral"], query["utm_medium"])
        self.assertEqual(["nejobs_pilot"], query["utm_campaign"])

    def test_approved_output_omits_closed_selected_vacancy(self):
        current = datetime.fromisoformat("2026-07-29T12:00:00+01:00")
        rows = approved_output_rows(
            [
                vacancy("open", closing_date="29/07/2026 23:59"),
                vacancy("closed", closing_date="29/07/2026 09:00"),
            ],
            decisions("open", "closed"),
            now=current,
        )

        self.assertEqual(["nejobs-open"], [row["job_id"] for row in rows])

    def test_composer_replaces_old_nejobs_and_preserves_jobg8(self):
        jobg8 = {
            "job_id": "jobg8-1",
            "title": "Office Administrator",
            "company": "Private Employer",
            "location": "Newcastle",
            "source": "JobG8",
        }
        old_external = {
            "job_id": "nejobs-old",
            "title": "Old Administrator",
            "company": "Old Council",
            "location": "Durham",
            "source": "NEJobs",
        }
        new_external = vacancy_to_published_job(vacancy("new"))
        expired_external = vacancy_to_published_job(
            vacancy("expired", closing_date="28/07/2026")
        )

        composed, counts = compose_rows(
            [old_external, jobg8],
            [new_external, expired_external],
            today=date(2026, 7, 29),
        )

        self.assertEqual(
            ["nejobs-new", "jobg8-1"],
            [row["job_id"] for row in composed],
        )
        self.assertEqual(
            {
                "jobg8_or_other": 1,
                "nejobs": 1,
                "expired_nejobs_skipped": 1,
                "duplicate_nejobs_skipped": 0,
                "total": 2,
            },
            counts,
        )

    def test_composer_suppresses_exact_cross_source_factual_duplicate(self):
        jobg8 = {
            "job_id": "jobg8-1",
            "title": "Administration Assistant",
            "company": "Example Council",
            "location": "Durham",
            "source": "JobG8",
        }
        external = vacancy_to_published_job(vacancy("300001"))

        composed, counts = compose_rows(
            [jobg8],
            [external],
            today=date(2026, 7, 29),
        )

        self.assertEqual([jobg8], composed)
        self.assertEqual(1, counts["duplicate_nejobs_skipped"])

    def test_composer_uses_exact_early_closing_time_on_current_day(self):
        external = vacancy_to_published_job(
            vacancy("early", closing_date="29/07/2026 09:00")
        )

        composed, counts = compose_rows(
            [],
            [external],
            today=date(2026, 7, 29),
            now=datetime.fromisoformat("2026-07-29T12:00:00+01:00"),
        )

        self.assertEqual([], composed)
        self.assertEqual(1, counts["expired_nejobs_skipped"])


if __name__ == "__main__":
    unittest.main(verbosity=2)

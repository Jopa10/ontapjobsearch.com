from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook

from scripts.export_daily_region_overview_xlsx import export, extract_tables


SAMPLE = """# Overview

## SITEWIDE RECONCILIATION
| Measure | Count |
|---|---:|
| Unique live jobs | 1,671 |

## JOBG8 FEED RECEIVED
| JobG8 classification | Jobs received | Ontap jobs |
|---|---:|---:|
| Administration | 920 | 548 |
| Total Ontap JobG8 jobs published today | 920 | 548 |

## PAGES
| Level | Page type | Area | Family | URL | Page count | Live jobs | In sitemap |
|---|---|---|---|---|---:|---:|---|
| Summary | All published/indexable URLs | Sitewide |  |  | 1,964 |  | Yes |
| Detail | Core | Sitewide |  | / | 1 |  | Yes |

## CITY OPPORTUNITIES
| Status | Town/city/locality | Region | All live jobs | Existing pages | Current routes |
|---|---|---|---:|---:|---|
| CREATE | Lincoln | Lincolnshire | 12 | 0 |  |

## LIVE
| Region | Service admin |
|---|---:|
| Berkshire | 29 |

## NOT LIVE
| Region | Service admin |
|---|---:|
| Bedfordshire | 1 / 1.0 / 0/8 |
"""


class DailyOverviewXlsxTests(unittest.TestCase):
    def test_extracts_the_five_owner_tables(self) -> None:
        tables = extract_tables(SAMPLE)
        self.assertEqual(list(tables), ["Sitewide", "JobG8 categories", "PAGES", "CITY OPPORTUNITIES", "LIVE", "NOT LIVE"])
        self.assertEqual(tables["Sitewide"][1], ["Unique live jobs", 1671])
        self.assertEqual(tables["JobG8 categories"][1], ["Administration", 920, 548])

    def test_exports_five_sheet_workbook(self) -> None:
        with tempfile.TemporaryDirectory() as directory:
            root = Path(directory)
            markdown = root / "overview.md"
            output = root / "overview.xlsx"
            markdown.write_text(SAMPLE, encoding="utf-8")
            export(markdown, output)
            workbook = load_workbook(output, data_only=True)
            self.assertEqual(workbook.sheetnames, ["Sitewide", "JobG8 categories", "PAGES", "CITY OPPORTUNITIES", "LIVE", "NOT LIVE"])
            self.assertEqual(workbook["LIVE"]["B2"].value, 29)
            self.assertEqual(workbook["JobG8 categories"]["C2"].value, 548)
            self.assertEqual(workbook["PAGES"]["F2"].value, 1964)
            self.assertEqual(workbook["CITY OPPORTUNITIES"]["D2"].value, 12)
            self.assertFalse(workbook["NOT LIVE"]["B2"].alignment.wrap_text)
            self.assertEqual(workbook["NOT LIVE"].row_dimensions[2].height, 20)
            workbook.close()


if __name__ == "__main__":
    unittest.main()

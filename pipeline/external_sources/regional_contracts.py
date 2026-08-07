"""Shared regional contracts for Ontap external-source processing.

This module deliberately reuses Ontap's two existing authorities:

* ``geo/geo_lookup.xlsx`` routes factual locations to Ontap regions.
* ``registers/region_category_slice_register.csv`` controls LIVE/CANDIDATE status.

It does not create or maintain a replacement central registry.
"""
from __future__ import annotations

import csv
import json
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Iterable

from openpyxl import load_workbook


CATEGORY_ADMIN_SERVICE = "admin_service"
PUBLISHABLE_STATUS = "LIVE"
KNOWN_SLICE_STATUSES = {"LIVE", "CANDIDATE", "RETIRED"}
UNUSABLE_GEO_CLUSTERS = {"", "unknown", "not specified"}
GEO_FALLBACK_FILENAME = "location_fallbacks.csv"

# Existing public-page roll-up already used by the daily admin/service pipeline.
_PUBLIC_REGION_ROLLUPS = {
    "north east county durham darlington hartlepool": "North East",
    "north east tyneside wearside northumberland": "North East",
}


@dataclass(frozen=True)
class SliceAuthority:
    region: str
    category: str
    status: str

    @property
    def may_publish(self) -> bool:
        return self.status == PUBLISHABLE_STATUS


@dataclass(frozen=True)
class GeographyResult:
    status: str
    region: str = ""
    cluster: str = ""
    evidence: str = ""
    lookup_key: str = ""


@dataclass
class DiscoveryRecord:
    source: str
    source_job_id: str
    canonical_url: str
    title: str = ""
    employer: str = ""
    location: str = ""
    postcode: str = ""
    salary_text: str = ""
    posted_date: str = ""
    closing_date: str = ""
    employment_type: str = ""
    description_excerpt: str = ""
    discovery_routes: list[dict[str, object]] = field(default_factory=list)

    def stable_key(self) -> str:
        source_id = clean(self.source_job_id)
        if source_id:
            return f"id:{source_id.casefold()}"
        url = canonical_url(self.canonical_url)
        if url:
            return f"url:{url.casefold()}"
        raise ValueError("source-wide discovery row has no stable source ID or URL")

    def add_provenance(self, *, route: str, query: str, page: int) -> None:
        item = {"route": clean(route), "query": clean(query), "page": int(page)}
        if item not in self.discovery_routes:
            self.discovery_routes.append(item)

    def provenance_json(self) -> str:
        return json.dumps(self.discovery_routes, ensure_ascii=False, sort_keys=True)


def clean(value: object) -> str:
    return re.sub(r"\s+", " ", str(value or "").replace("\xa0", " ")).strip()


def normalise(value: object) -> str:
    return re.sub(r"[^a-z0-9]+", " ", clean(value).casefold()).strip()


def canonical_url(value: object) -> str:
    text = clean(value)
    return text.split("#", 1)[0].rstrip("/")


def canonical_public_region(cluster: object) -> str:
    """Return the existing public slice region for one lookup cluster."""
    value = clean(cluster)
    key = normalise(value)
    if key in UNUSABLE_GEO_CLUSTERS:
        return ""
    if key in _PUBLIC_REGION_ROLLUPS:
        return _PUBLIC_REGION_ROLLUPS[key]
    if key == "devon":
        return "Devon"
    return value


def load_slice_authorities(path: Path) -> dict[tuple[str, str], SliceAuthority]:
    if not path.is_file():
        raise ValueError(f"slice register not found: {path}")
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        if reader.fieldnames != ["region", "category", "status"]:
            raise ValueError(
                "slice register columns must be exactly region,category,status"
            )
        output: dict[tuple[str, str], SliceAuthority] = {}
        for line_number, row in enumerate(reader, start=2):
            region = clean(row.get("region"))
            category = clean(row.get("category"))
            status = clean(row.get("status")).upper()
            if not region or not category or status not in KNOWN_SLICE_STATUSES:
                raise ValueError(f"invalid slice register row at line {line_number}")
            key = (region, category)
            if key in output:
                raise ValueError(f"duplicate slice authority: {region} / {category}")
            output[key] = SliceAuthority(region, category, status)
    return output


def slice_authority(
    authorities: dict[tuple[str, str], SliceAuthority],
    *,
    region: str,
    category: str,
) -> SliceAuthority | None:
    """Return the explicit authority; absence never implies LIVE."""
    return authorities.get((clean(region), clean(category)))


def publishable_region(
    authorities: dict[tuple[str, str], SliceAuthority],
    *,
    region: str,
    category: str,
) -> bool:
    authority = slice_authority(authorities, region=region, category=category)
    return bool(authority and authority.may_publish)


def merge_discovery_records(records: Iterable[DiscoveryRecord]) -> list[DiscoveryRecord]:
    """Deduplicate source-wide discovery before geography or classification."""
    merged: dict[str, DiscoveryRecord] = {}
    for record in records:
        key = record.stable_key()
        current = merged.get(key)
        if current is None:
            merged[key] = record
            continue
        factual_fields = (
            "source",
            "source_job_id",
            "canonical_url",
            "title",
            "employer",
            "location",
            "postcode",
            "salary_text",
            "posted_date",
            "closing_date",
            "employment_type",
            "description_excerpt",
        )
        for field_name in factual_fields:
            old = clean(getattr(current, field_name))
            new = clean(getattr(record, field_name))
            if field_name == "canonical_url":
                old = canonical_url(old)
                new = canonical_url(new)
            if old and new and old != new:
                raise ValueError(
                    f"conflicting factual discovery values for {key}: {field_name}"
                )
            if not old and new:
                setattr(current, field_name, new)
        for item in record.discovery_routes:
            current.add_provenance(
                route=clean(item.get("route")),
                query=clean(item.get("query")),
                page=int(item.get("page") or 1),
            )
    return sorted(merged.values(), key=lambda item: item.stable_key())


_GEO_COLUMN_ALIASES = {
    "lookup": (
        "location",
        "location_clean",
        "lookup",
        "lookup_value",
        "area",
        "town",
        "place",
        "postcode",
        "postcode_area",
    ),
    "region": (
        "region",
        "ontap_region",
        "region_cluster",
        "cluster",
    ),
}


def _column_index(headers: list[str], aliases: tuple[str, ...]) -> int | None:
    normalised = [normalise(header).replace(" ", "_") for header in headers]
    for alias in aliases:
        target = normalise(alias).replace(" ", "_")
        if target in normalised:
            return normalised.index(target)
    return None


def _add_geo_mapping(
    output: dict[str, str],
    *,
    lookup_value: object,
    cluster: object,
    source: str,
) -> None:
    key = normalise(lookup_value)
    region_cluster = clean(cluster)
    if not key or not canonical_public_region(region_cluster):
        return
    existing = output.get(key)
    if existing and normalise(existing) != normalise(region_cluster):
        raise ValueError(
            f"ambiguous geographic lookup value {key!r}: "
            f"{existing!r} versus {region_cluster!r} ({source})"
        )
    if not existing:
        output[key] = region_cluster


def _load_geo_sheet(
    worksheet: object,
    *,
    output: dict[str, str],
    source: str,
    require_auto_status: bool = False,
) -> None:
    rows = worksheet.iter_rows(values_only=True)
    try:
        headers = [clean(value) for value in next(rows)]
    except StopIteration as exc:
        raise ValueError(f"geographic lookup sheet is empty: {source}") from exc
    lookup_index = _column_index(headers, _GEO_COLUMN_ALIASES["lookup"])
    region_index = _column_index(headers, _GEO_COLUMN_ALIASES["region"])
    status_index = _column_index(headers, ("status",))
    if lookup_index is None or region_index is None:
        raise ValueError(
            f"geographic lookup sheet {source} must expose a recognised "
            "location and cluster column"
        )
    if require_auto_status and status_index is None:
        raise ValueError(
            f"geographic lookup sheet {source} must expose a Status column"
        )
    for row_number, row in enumerate(rows, start=2):
        if require_auto_status:
            status = clean(
                row[status_index] if status_index is not None and status_index < len(row) else ""
            )
            if normalise(status) != "auto":
                continue
        lookup_value = row[lookup_index] if lookup_index < len(row) else ""
        cluster = row[region_index] if region_index < len(row) else ""
        _add_geo_mapping(
            output,
            lookup_value=lookup_value,
            cluster=cluster,
            source=f"{source} row {row_number}",
        )


def _load_geo_fallback_csv(
    path: Path,
    *,
    output: dict[str, str],
) -> None:
    """Load shared factual fallbacks without embedding source-specific rules."""
    if not path.is_file():
        return
    with path.open(newline="", encoding="utf-8-sig") as handle:
        reader = csv.DictReader(handle)
        expected = ["lookup_value", "region", "status"]
        if reader.fieldnames != expected:
            raise ValueError(
                f"geographic fallback columns must be exactly {','.join(expected)}"
            )
        for line_number, row in enumerate(reader, start=2):
            status = clean(row.get("status"))
            if normalise(status) != "auto":
                continue
            lookup_value = clean(row.get("lookup_value"))
            region = clean(row.get("region"))
            if not lookup_value or not region:
                raise ValueError(
                    f"invalid geographic fallback row at line {line_number}"
                )
            _add_geo_mapping(
                output,
                lookup_value=lookup_value,
                cluster=region,
                source=f"{path.name} row {line_number}",
            )


def load_geo_lookup(path: Path) -> dict[str, str]:
    """Load Ontap's Area/Cluster and approved location-fallback contracts."""
    if not path.is_file():
        raise ValueError(f"geographic lookup not found: {path}")
    workbook = load_workbook(path, read_only=True, data_only=True)
    output: dict[str, str] = {}
    _load_geo_sheet(
        workbook.active,
        output=output,
        source=workbook.active.title,
    )
    if "LocationFallback" in workbook.sheetnames:
        _load_geo_sheet(
            workbook["LocationFallback"],
            output=output,
            source="LocationFallback",
            require_auto_status=True,
        )
    _load_geo_fallback_csv(
        path.with_name(GEO_FALLBACK_FILENAME),
        output=output,
    )
    if not output:
        raise ValueError("geographic lookup contains no usable mappings")
    return output


def _postcode_tiers(postcode: str) -> list[tuple[str, tuple[str, ...]]]:
    key = normalise(postcode)
    if not key:
        return []
    tiers: list[tuple[str, tuple[str, ...]]] = [("postcode", (key,))]
    outward = re.match(r"^([a-z]{1,2}\d[a-z\d]?)\b", key)
    if outward:
        tiers.append(("postcode outward", (outward.group(1),)))
    area = re.match(r"^([a-z]{1,2})\d", key)
    if area:
        tiers.append(("postcode area", (area.group(1),)))
    return tiers


def _location_tiers(location: str) -> list[tuple[str, tuple[str, ...]]]:
    raw = clean(location)
    if not raw:
        return []
    full = normalise(raw)
    parts = [normalise(part) for part in raw.split(",") if normalise(part)]
    tiers: list[tuple[str, tuple[str, ...]]] = []
    if full:
        tiers.append(("full location", (full,)))
    if parts:
        tiers.append(("locality", (parts[0],)))
    for index, part in enumerate(parts[1:], start=2):
        tiers.append((f"location component {index}", (part,)))
    return tiers


def route_geography(
    *,
    location: str,
    postcode: str,
    lookup: dict[str, str],
) -> GeographyResult:
    """Route factual location evidence only; uncertain geography remains visible.

    Evidence is evaluated from most to least specific. A factual locality match
    therefore wins before a broader county component, while conflicting matches
    at the same evidence tier remain unresolved.
    """
    tiers = _postcode_tiers(postcode) + _location_tiers(location)
    seen_keys: set[str] = set()
    for evidence_type, raw_keys in tiers:
        keys = tuple(key for key in raw_keys if key and key not in seen_keys)
        seen_keys.update(keys)
        matches: dict[str, list[tuple[str, str]]] = {}
        for key in keys:
            cluster = lookup.get(key)
            region = canonical_public_region(cluster)
            if not region:
                continue
            matches.setdefault(region, []).append((clean(cluster), key))
        if not matches:
            continue
        if len(matches) > 1:
            return GeographyResult(
                status="UNRESOLVED",
                evidence=(
                    f"Conflicting exact {evidence_type} matches in geo_lookup.xlsx: "
                    + ", ".join(sorted(matches))
                ),
            )
        region, evidence_rows = next(iter(matches.items()))
        cluster, key = evidence_rows[0]
        return GeographyResult(
            status="ROUTED",
            region=region,
            cluster=cluster,
            evidence=f"Exact {evidence_type} match in geo_lookup.xlsx",
            lookup_key=key,
        )
    return GeographyResult(
        status="UNRESOLVED",
        evidence="No exact factual location match in geo_lookup.xlsx",
    )
